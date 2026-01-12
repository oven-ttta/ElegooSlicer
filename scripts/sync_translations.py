#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Translation Sync Script for ElegooSlicer
"""

import os
from pathlib import Path
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class TranslationSyncer:
    def __init__(self, elegoo_i18n_path, orca_i18n_path):
        self.elegoo_i18n_path = Path(elegoo_i18n_path)
        self.orca_i18n_path = Path(orca_i18n_path)
        
    def find_po_files(self):
        """Find all ElegooSlicer PO files"""
        po_files = []
        for po_file in self.elegoo_i18n_path.glob("*/ElegooSlicer_*.po"):
            po_files.append(po_file)
        return po_files
    
    def parse_elegoo_po_file(self, file_path):
        """Parse ElegooSlicer PO file - only get msgid that need translation (empty msgstr)"""
        translations = {}
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # Pre-filter: remove empty lines and comment lines
            lines = [line.strip() for line in lines if line.strip() and not line.strip().startswith('#~')]
            
            i = 0
            while i < len(lines):
                line = lines[i]
                
                if line.startswith('msgid '):
                    # Read msgid content
                    msgid_content = line[line.find('"'):].strip()
                    i += 1
                    
                    # Read multi-line msgid
                    while i < len(lines) and lines[i].startswith('"'):
                        msgid_content += "\n" + lines[i]
                        i += 1
                    # Skip empty msgid
                    if msgid_content == '""':
                        continue
                    
                    # Read msgstr content
                    if i < len(lines) and lines[i].startswith('msgstr '):
                        msgstr_content = lines[i][lines[i].find('"'):].strip()
                        i += 1
                        
                        # Read multi-line msgstr
                        while i < len(lines) and lines[i].startswith('"'):
                            msgstr_content += "\n" + lines[i]
                            i += 1
                        
                        # Save if msgstr is empty
                        if msgstr_content == '""' or not msgstr_content:
                            translations[msgid_content] = ""
                    else:
                        # No msgstr found
                        translations[msgid_content] = ""
                else:
                    i += 1
                    
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
            
        return translations
    
    def parse_elegoo_po_file_all(self, file_path):
        """Parse ElegooSlicer PO file - get ALL msgid->msgstr (only non-empty msgstr)."""
        translations = {}
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # Pre-filter: remove empty lines and comment lines
            lines = [line.strip() for line in lines if line.strip() and not line.strip().startswith('#~')]
            
            i = 0
            while i < len(lines):
                line = lines[i]
                
                if line.startswith('msgid '):
                    # Read msgid content
                    msgid_content = line[line.find('"'):].strip()
                    i += 1
                    
                    # Read multi-line msgid
                    while i < len(lines) and lines[i].startswith('"'):
                        msgid_content += "\n" + lines[i]
                        i += 1
                    
                    # Skip empty msgid
                    if msgid_content == '""':
                        continue
                    
                    # Read msgstr content
                    if i < len(lines) and lines[i].startswith('msgstr '):
                        msgstr_content = lines[i][lines[i].find('"'):].strip()
                        i += 1
                        
                        # Read multi-line msgstr
                        while i < len(lines) and lines[i].startswith('"'):
                            msgstr_content += "\n" + lines[i]
                            i += 1
                        
                        # Save only if msgstr is non-empty
                        if msgstr_content and msgstr_content != '""':
                            translations[msgid_content] = msgstr_content
                    else:
                        # No msgstr found - skip this entry
                        continue
                else:
                    i += 1
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
        
        return translations
    
    def parse_orca_po_file(self, file_path):
        """Parse OrcaSlicer PO file - get all translations for lookup, including comment lines"""
        translations = {}
        comment_lines = {}  # Store comment lines (like #, fuzzy, c-format) for each msgid
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            i = 0
            while i < len(lines):
                line_stripped = lines[i].strip()
                
                # Look for comment lines before msgid (like #, fuzzy, c-format)
                comment_line = None
                if line_stripped.startswith('#,') and not line_stripped.startswith('#~'):
                    comment_line = lines[i]  # Keep original format with newline
                    i += 1
                    # Skip empty lines after comment
                    while i < len(lines) and not lines[i].strip():
                        i += 1
                
                if i < len(lines) and lines[i].strip().startswith('msgid '):
                    # Read msgid content
                    msgid_line = lines[i].strip()
                    msgid_content = msgid_line[msgid_line.find('"'):].strip()
                    i += 1
                    
                    # Read multi-line msgid
                    while i < len(lines) and (lines[i].strip().startswith('"') or not lines[i].strip()):
                        if lines[i].strip() and not lines[i].strip().startswith('#~'):
                            msgid_content += "\n" + lines[i].strip()
                        i += 1
                    
                    # Skip empty msgid
                    if msgid_content == '""':
                        continue

                    # Read msgstr content
                    if i < len(lines) and lines[i].strip().startswith('msgstr '):
                        msgstr_content = lines[i][lines[i].find('"'):].strip()
                        i += 1
                        
                        # Read multi-line msgstr
                        while i < len(lines) and (lines[i].strip().startswith('"') or not lines[i].strip()):
                            if lines[i].strip() and not lines[i].strip().startswith('#~'):
                                msgstr_content += "\n" + lines[i].strip()
                            i += 1
                        
                        # Save translation and comment line
                        if msgstr_content == '""' or not msgstr_content or 'orca' in msgstr_content.lower():
                            continue
                        translations[msgid_content] = msgstr_content
                        if comment_line:
                            comment_lines[msgid_content] = comment_line
                    else:
                        # No msgstr found - skip this entry
                        i += 1
                else:
                    i += 1
                    
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
            
        return translations, comment_lines
    
    def update_elegoo_po_file(self, po_file_path, elegoo_translations, orca_translations, orca_comment_lines):
        """Update ElegooSlicer PO file with translations from OrcaSlicer, including comment lines"""
        updated_count = 0
        changes = []  # 记录修改的详细信息
        
        try:
            # Read all lines first - keep original format for writing back
            with open(po_file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # Process each line and update in place
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                # Skip empty lines and obsolete comments
                if not line or line.startswith('#~'):
                    i += 1
                    continue
                
                # Look for comment line before msgid (like #, fuzzy, c-format)
                comment_line_idx = None
                if line.startswith('#,'):
                    comment_line_idx = i
                    i += 1
                    # Skip empty lines after comment
                    while i < len(lines) and not lines[i].strip():
                        i += 1
                    if i < len(lines):
                        line = lines[i].strip()
                
                if line.startswith('msgid '):
                    # Read msgid content
                    first_quote = line.find('"')
                    if first_quote != -1:
                        msgid_content = line[first_quote:].strip()
                        
                        # Read multi-line msgid
                        j = i + 1
                        while j < len(lines) and (lines[j].strip().startswith('"') or not lines[j].strip()):
                            if lines[j].strip() and not lines[j].strip().startswith('#~'):
                                msgid_content += "\n" + lines[j].strip()
                            j += 1
                        
                        # Skip empty msgid (file header)
                        if msgid_content == '""':
                            i = j
                            continue
                        
                        # Skip empty lines between msgid and msgstr
                        while j < len(lines) and (not lines[j].strip() or lines[j].strip().startswith('#~')):
                            j += 1
                        
                        # Look for msgstr
                        if j < len(lines) and lines[j].strip().startswith('msgstr '):
                            # Check if we need to update this msgid
                            if msgid_content in elegoo_translations and msgid_content in orca_translations:
                                orca_msgstr = orca_translations[msgid_content]
                                
                                # Skip if OrcaSlicer msgstr is empty
                                if orca_msgstr.strip():
                                    # Skip if OrcaSlicer translation contains "Orca"
                                    if 'orca' in orca_msgstr.lower():
                                        print(f"    Warning: Skipping translation containing 'Orca': {orca_msgstr}")
                                        i = j
                                        continue
                                    
                                    # Record the change
                                    changes.append({
                                        'msgid': msgid_content,
                                        'translation': orca_msgstr
                                    })
                                    
                                    # Update or add/remove comment line based on OrcaSlicer
                                    if msgid_content in orca_comment_lines:
                                        # OrcaSlicer has comment line - sync it
                                        orca_comment = orca_comment_lines[msgid_content]
                                        if comment_line_idx is not None:
                                            # Update existing comment line
                                            lines[comment_line_idx] = orca_comment
                                        else:
                                            # Add new comment line before msgid
                                            lines.insert(i, orca_comment)
                                            j += 1  # Adjust j since we inserted a line
                                    else:
                                        # OrcaSlicer has no comment line - remove ElegooSlicer's if exists
                                        if comment_line_idx is not None:
                                            # Remove the comment line
                                            del lines[comment_line_idx]
                                            # Adjust indices since we removed a line
                                            if comment_line_idx < i:
                                                i -= 1
                                            if comment_line_idx < j:
                                                j -= 1
                                    
                                    # Update the msgstr line in original file
                                    lines[j] = f'msgstr {orca_msgstr}\n'
                                    updated_count += 1
                                    
                                    # Skip to next msgid
                                    i = j + 1
                                    continue
                            
                            # No update needed, continue to next msgid
                            i = j + 1
                        else:
                            # No msgstr found, continue to next msgid
                            i = j
                    else:
                        i += 1
                else:
                    i += 1
            
            # Write back the updated content with original format preserved
            with open(po_file_path, 'w', encoding='utf-8') as f:
                f.writelines(lines)
                
        except Exception as e:
            print(f"Error updating {po_file_path}: {e}")
            
        return updated_count, changes
    

    
    def sync_translations(self):
        """Main method to sync all translations"""
        print("Starting translation sync...")
        print(f"ElegooSlicer path: {self.elegoo_i18n_path}")
        print(f"OrcaSlicer path: {self.orca_i18n_path}")
        
        po_files = self.find_po_files()
        print(f"Found {len(po_files)} PO files to process")
        
        total_updated = 0
        all_changes = {}  # 记录所有文件的修改信息
        all_need_translation = {}  # 记录所有需要翻译的条目
        all_translated = {}  # 记录已翻译条目（从 ElegooSlicer）
        
        for i, po_file in enumerate(po_files, 1):
            language_code = po_file.parent.name
            print(f"\n[{i}/{len(po_files)}] Processing: {po_file.absolute()}")
            
            # Parse ElegooSlicer PO file - only get msgid that need translation
            elegoo_translations = self.parse_elegoo_po_file(po_file)
            # Parse ALL ElegooSlicer entries that are already translated
            elegoo_translated = self.parse_elegoo_po_file_all(po_file)

            print(f"  ElegooSlicer: {len(elegoo_translations)} translations need updating, {len(elegoo_translated)} already translated")
            
            # Store entries for reporting regardless of whether updates are needed
            all_translated[language_code] = elegoo_translated
            
            # Store all entries that need translation
            all_need_translation[language_code] = elegoo_translations
            
            # Find corresponding OrcaSlicer translations
            orca_po_file = self.orca_i18n_path / language_code / f"OrcaSlicer_{language_code}.po"
            print(f"  OrcaSlicer file: {orca_po_file.absolute()}")
            
            if not orca_po_file.exists():
                print(f"  OrcaSlicer file not found: {orca_po_file.absolute()}")
                continue
            
            # Parse OrcaSlicer PO file
            orca_translations, orca_comment_lines = self.parse_orca_po_file(orca_po_file)
            print(f"  OrcaSlicer: {len(orca_translations)} translations")
            
            # Count translations that can be updated
            need_update = 0
            found_in_orca = 0
            empty_orca = 0
            contains_orca = 0
            
            # print(f"  ElegooSlicer msgids that need updating:")
            # for msgid in elegoo_translations:
            #     print(f"    - {msgid}")
            
            print(f"  Checking OrcaSlicer translations:")

            for msgid in elegoo_translations:
                if msgid in orca_translations:
                    found_in_orca += 1
                    orca_msgstr = orca_translations[msgid]
                    print(f"    Found: '{msgid}' -> '{orca_msgstr}'")
                    # Only count if OrcaSlicer has non-empty translation and doesn't contain "Orca"
                    if orca_msgstr.strip():
                        if 'orca' not in orca_msgstr.lower():
                            need_update += 1
                        else:
                            contains_orca += 1
                    else:
                        empty_orca += 1
                else:
                    print(f"    NOT FOUND: '{msgid}'")
            
            print(f"  Found in OrcaSlicer: {found_in_orca}")
            print(f"  OrcaSlicer empty: {empty_orca}")
            print(f"  Contains 'Orca': {contains_orca}")
            print(f"  Can update: {need_update} translations")
            
            # Update translations
            updated_count, changes = self.update_elegoo_po_file(po_file, elegoo_translations, orca_translations, orca_comment_lines)
            total_updated += updated_count
            print(f"  Actually updated: {updated_count} translations")
            
            # Store changes for this file
            if changes:
                all_changes[language_code] = changes
        
        print(f"\nSync completed! Total translations updated: {total_updated}")
        
        # Generate Excel report
        if all_need_translation or all_translated:
            self.generate_excel_report(all_need_translation, all_changes, all_translated)
            print(f"Excel report generated: translation_changes.xlsx")
    
    def generate_excel_report(self, all_need_translation, all_changes, all_translated):
        """Generate Excel report with translation changes"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Ensure we iterate over all languages present in either dict
        language_codes = set(all_need_translation.keys()) | set(all_translated.keys())
        for language_code in sorted(language_codes):
            translations_to_update = all_need_translation.get(language_code, {})
            # Create worksheet for each language
            ws = wb.create_sheet(title=language_code)
            
            # Set column headers
            ws['A1'] = 'msgid'
            ws['B1'] = 'Translation'
            
            # Apply header style
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['B1'].font = header_font
            ws['B1'].fill = header_fill
            
            # Get changes for this language (if any)
            changes = all_changes.get(language_code, {})
            
            # Add data - sort by status: updated first, then untranslated
            updated_entries = []
            untranslated_entries = []
            
            for msgid in translations_to_update:
                # Check if this msgid was updated
                was_updated = False
                orca_translation = ""
                for change in changes:
                    if change['msgid'] == msgid:
                        orca_translation = change['translation']
                        was_updated = True
                        break
                
                entry = {
                    'msgid': msgid,
                    'translation': orca_translation if was_updated else ""
                }
                
                if was_updated:
                    updated_entries.append(entry)
                else:
                    untranslated_entries.append(entry)
            
            # Sort: updated first, then untranslated
            sorted_entries = updated_entries + untranslated_entries
            
            # Write sorted data to Excel
            for i, entry in enumerate(sorted_entries, 2):
                ws[f'A{i}'] = entry['msgid']
                ws[f'B{i}'] = entry['translation']
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 80
            ws.column_dimensions['B'].width = 80

            # Also export already translated entries into a second sheet
            translated_entries = all_translated.get(language_code, {})
            ws_t = wb.create_sheet(title=f"{language_code}_translated")
            ws_t['A1'] = 'msgid'
            ws_t['B1'] = 'Elegoo Translation'
            ws_t['A1'].font = header_font
            ws_t['A1'].fill = header_fill
            ws_t['B1'].font = header_font
            ws_t['B1'].fill = header_fill

            for i, (msgid, translation) in enumerate(sorted(translated_entries.items()), 2):
                ws_t[f'A{i}'] = msgid
                ws_t[f'B{i}'] = translation
            ws_t.column_dimensions['A'].width = 80
            ws_t.column_dimensions['B'].width = 80
        
        # Save the workbook
        wb.save('translation_changes.xlsx')


def main():
    parser = argparse.ArgumentParser(description='Sync translations from OrcaSlicer to ElegooSlicer')
    parser.add_argument('--elegoo-path', default='../localization/i18n', 
                       help='Path to ElegooSlicer i18n directory (default: current directory)')
    parser.add_argument('--orca-path', default='../../SoftFever/OrcaSlicer/localization/i18n',
                       help='Path to OrcaSlicer i18n directory')
    
    args = parser.parse_args()
    
    elegoo_path = Path(args.elegoo_path)
    orca_path = Path(args.orca_path)
    
    if not elegoo_path.exists():
        print(f"Error: ElegooSlicer path does not exist: {elegoo_path}")
        return
        
    if not orca_path.exists():
        print(f"Error: OrcaSlicer path does not exist: {orca_path}")
        return
    
    syncer = TranslationSyncer(elegoo_path, orca_path)
    syncer.sync_translations()


if __name__ == "__main__":
    main()
